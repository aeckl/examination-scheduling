import openpyxl as op
import numpy as np
import random as rd

def load_rooms():
	#get woorkbook and worksheet
    wb = op.load_workbook('model\Data\Raumuebersicht.xlsx')
    sheet = wb.get_sheet_by_name('Raum')

    #load rooms, capayity for exams and id of the campus; 01 = Innenstadt, 02 = Garching, 02-81 = Hochbrueck
    rooms = []
    seatings = {}
    id_campus = {}
    for i in range(2,sheet.max_row+1):
        rooms.extend([sheet.cell(row=i, column=1).value])
        seatings.update({"%s" % (sheet.cell(row=i, column=1).value)  :  int(sheet.cell(row=i, column=4).value) } )
        id_campus.update({"%s" % (sheet.cell(row=i, column=1).value) :  "%s" % (sheet.cell(row=i, column=9).value) } ) 


    return rooms, seatings, id_campus

def get_random_room_capacity(r,w):
    """
    r = number of rooms 
    w = where (01    = Innenstadt, 
               02    = Garching, 
               02-81 = Hochbrueck) 

    """
    #Load rooms, seating capacicity and id of campus
    rooms, seatings, id_campus = load_rooms()


    #Generate room capacity by from rooms selected randomly by id
    room_capacity = [seatings[rooms[i]] for i in np.random.choice([i for i in range(len(rooms)) if id_campus[rooms[i]] in w],r)]


    return room_capacity




#build_smart_random(tseed=500, n=100,r=20,p=20)